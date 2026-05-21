from openpyxl import load_workbook
import datetime as dt
from bs4 import BeautifulSoup
import requests

# Creating workbook/worksheet to work on
wb = load_workbook("5050 lottery data.xlsx")
ws = wb.active

if ws.title != dt.datetime.now().strftime("%b"):
    ws.create_sheet(dt.datetime.now().strftime("%b"))
else:
    ws = wb[dt.datetime.now().strftime("%b")]
wb.save("5050 lottery data.xlsx")

# Adding values in worksheet
ws["A1"] = "Date"
ws["B1"] = "HSN Lottery"
ws["C1"] = "splitthepot"
ws["D1"] = "ThunderBay5050"
headers = {"User-Agent": "Mozilla/5.0"}

def job():
    cellno = ws.max_row + 1
    ws[f"A{cellno}"] = dt.datetime.now().strftime("%Y-%m-%d")

    url = "https://hsn5050.ca/?gad_source=1&gad_campaignid=23815246966&gbraid=0AAAAA9mwGnJAFwpDZkWcbhho7dHGO5X0z&gclid=CjwKCAjw2rrQBhBuEiwAarLWHYXAde_qipIH4FC--FrEtQvSMyQPhMTkAF806z2bCWdGmDixuJ_uNBoCnNAQAvD_BwE"
    response = requests.get(url, headers=headers)
    soup = BeautifulSoup(response.content, 'lxml')
    prize = soup.find("span", class_ = "font-bold").text.strip()
    ws[f"B{cellno}"] = prize


    url1 = "https://on.splitthepot.ca/"
    response1 = requests.get(url1, headers=headers)
    soup1 = BeautifulSoup(response1.content, 'lxml')
    prize1 = soup1.find("span", class_ = "font-bold").text.strip()
    ws[f"C{cellno}"] = prize1

    url1 = "https://www.thunderbay5050.ca/"
    response1 = requests.get(url1, headers=headers)
    soup1 = BeautifulSoup(response1.content, 'lxml')
    prize2 = soup1.find("div", class_ = "draw-total").text.strip()
    ws[f"D{cellno}"] = prize2

job()
wb.save("5050 lottery data.xlsx")


